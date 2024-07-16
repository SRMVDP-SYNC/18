from flask import Flask, render_template, request, jsonify, send_file
import os
import pdfplumber
import pandas as pd
import json
from openpyxl import Workbook

app = Flask(__name__)

# Function to extract text from PDF
def extract_text_from_pdf(pdf_path):
    text_lines = []
    with pdfplumber.open(pdf_path) as pdf:
        for page_num, page in enumerate(pdf.pages):
            text = page.extract_text()
            if text:
                lines = text.split('\n')
                text_lines.extend(lines)
    return text_lines

# Function to extract information from text
def extract_info_from_text(text_lines):
    info = {
        "Program Section:": "",
        "Subject Code & Title:": "",
        "Test Name:": ""
    }
    for line in text_lines:
        for key in info.keys():
            if line.startswith(key):
                info[key] = line
    return info

# Function to extract marks from text
def extract_marks_from_text(text_lines, reg_numbers):
    marks_dict = {}
    for line in text_lines:
        for reg_number in reg_numbers:
            if reg_number in line:
                parts = line.split()
                reg_index = parts.index(reg_number)
                if reg_index + 1 < len(parts):
                    marks_dict[reg_number] = parts[reg_index + 1]
    return marks_dict

# Function to process PDFs in a folder
def process_pdfs_in_folder(file_order, folder_path, reg_numbers):
    results = {}
    info = {
        "Program Section:": "",
        "Subject Code & Title:": "",
        "Test Name:": ""
    }
    for filename in file_order:
        pdf_path = os.path.join(folder_path, filename)
        text_lines = extract_text_from_pdf(pdf_path)
        marks = extract_marks_from_text(text_lines, reg_numbers)
        for reg_number, marks_value in marks.items():
            if reg_number not in results:
                results[reg_number] = {}
            results[reg_number][filename] = marks_value
        extracted_info = extract_info_from_text(text_lines)
        for key in info:
            if extracted_info[key]:
                info[key] = extracted_info[key]
    return results, info

# Function to create DataFrame for component split-up
def input_split_up(components, array_2d):
    split_up_df = pd.DataFrame(columns=['Component', 'CO1', 'CO2', 'CO3', 'CO4', 'CO5', 'CO6'])
    split_up_df['Component'] = components

    for index, row in split_up_df.iterrows():
        for i, co in enumerate(['CO1', 'CO2', 'CO3', 'CO4', 'CO5', 'CO6']):
            split_up_df.at[index, co] = array_2d[index][i]

    return split_up_df

# Function to calculate total possible attainment
def calculate_total_possible_attainment(split_up_df):
    total_attainment = split_up_df[['CO1', 'CO2', 'CO3', 'CO4', 'CO5', 'CO6']].sum()
    return total_attainment

# Function to calculate row-wise sum
def calculate_row_wise_sum(split_up_df):
    split_up_df['Row-wise Sum'] = split_up_df.iloc[:, 1:].sum(axis=1)
    return split_up_df

# Function to calculate attainment
def calculate_attainment(split_up_df, marks_df, reg_numbers):
    attainment_df = pd.DataFrame(columns=['Register number', 'CO1', 'CO2', 'CO3', 'CO4', 'CO5', 'CO6'])
    attainment_df['Register number'] = reg_numbers

    for index, reg_number in enumerate(reg_numbers):
        attainment_values = {co: 0.0 for co in ['CO1', 'CO2', 'CO3', 'CO4', 'CO5', 'CO6']}
        for component in marks_df.columns:
            if component in split_up_df['Component'].values:
                student_marks = marks_df.loc[reg_number, component]
                row_sum = split_up_df.loc[split_up_df['Component'] == component, 'Row-wise Sum'].values[0]
                try:
                    student_marks = float(student_marks)
                    percentage_scored = (student_marks / row_sum) * 100
                except ValueError:
                    percentage_scored = 0

                for co in ['CO1', 'CO2', 'CO3', 'CO4', 'CO5', 'CO6']:
                    split_value = split_up_df.loc[split_up_df['Component'] == component, co].values[0]
                    attainment_values[co] += (percentage_scored * split_value) / 100

        for co in ['CO1', 'CO2', 'CO3', 'CO4', 'CO5', 'CO6']:
            attainment_df.at[index, co] = attainment_values[co]

    return attainment_df

# Function to write to Excel
def write_to_excel(df_transposed, split_up_df_without_sum, attainment_table, total_possible_attainment, target_value, output_path, info, file_order):
    output_path = os.path.join(os.path.dirname(output_path), 'results.xlsx')

    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        df_transposed.to_excel(writer, sheet_name='marks', index=True, float_format="%.2f")
        marks_sheet = writer.sheets['marks']
        marks_sheet.column_dimensions['A'].width = 20
        for col in ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']:
            marks_sheet.column_dimensions[col].width = 15
        split_up_df_without_sum.to_excel(writer, sheet_name='attainment_table', index=False, startrow=6, float_format="%.2f")
        attainment_table.to_excel(writer, sheet_name='attainment_table', index=False, startrow=split_up_df_without_sum.shape[0] + 8, float_format="%.2f")
        wb = writer.book
        attainment_sheet = wb['attainment_table']
        # Formatting and data entry
        attainment_sheet.column_dimensions['A'].width = 20
        for col in ['B', 'C', 'D', 'E', 'F', 'G', 'H']:
            attainment_sheet.column_dimensions[col].width = 10

        # Additional formatting and data entry
        attainment_sheet.insert_rows(1, 6)
        attainment_sheet.merge_cells('H1:M1')
        attainment_sheet['H1'] = 'SRM Institute of Science and Technology Vadapalani'
        attainment_sheet.row_dimensions[1].height = 30

        attainment_sheet.merge_cells('H2:M2')
        attainment_sheet['H2'] = info['Program Section:']
        attainment_sheet.row_dimensions[2].height = 30

        attainment_sheet.merge_cells('H3:M3')
        attainment_sheet['H3'] = info['Subject Code & Title:']
        attainment_sheet.row_dimensions[3].height = 30

        attainment_sheet.merge_cells('H4:M4')
        attainment_sheet['H4'] = info['Test Name:']
        attainment_sheet.row_dimensions[4].height = 30

        # Write total possible attainment row
        total_row = ['Total Possible Attainment'] + total_possible_attainment.tolist()
        attainment_sheet.append(total_row)

        # Write target (75% of total possible CO attainment)
        attainment_sheet.append(["Target (75% of total possible CO attainment)"] + [f"{value:.2f}" for value in target_value])

        # Calculate attainment levels for each CO
        num_students_above_target = (attainment_table.iloc[:, 1:] > target_value).sum()
        attainment_sheet.append(["No of students > Target"] + num_students_above_target.tolist())

        percent_students_above_target = (num_students_above_target / len(attainment_table)) * 100
        attainment_sheet.append(["% of students > Target"] + [f"{value:.2f}%" for value in percent_students_above_target])

        # Calculate attainment levels based on percent_students_above_target for each CO
        attainment_levels = []

        for co in ['CO1', 'CO2', 'CO3', 'CO4', 'CO5', 'CO6']:
            percentage = percent_students_above_target[co]
            if percentage >= 80:
                attainment_levels.append('3')
            elif 70 <= percentage < 80:
                attainment_levels.append('2')
            elif 60 <= percentage < 70:
                attainment_levels.append('1')
            else:
                attainment_levels.append('0')

        # Append Attainment Levels for each CO
        attainment_labels = ["Attainment Level  %>=80 -- 3   70<=%<80  --2  60<=%<70  --1   %<60 -- 0"]
        attainment_levels_row = attainment_labels + attainment_levels
        attainment_sheet.append(attainment_levels_row)

        # Calculate average of all CO attainment levels
        avg_attainment_levels = pd.Series([int(level) for level in attainment_levels]).mean()
        attainment_sheet.append(["Average of all CO attainment levels"] + [f"{avg_attainment_levels:.2f}"])

        # Replace .pdf extensions in the data with empty string
        for row in attainment_sheet.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.endswith('.pdf'):
                    cell.value = cell.value[:-4]
                    
        for row in marks_sheet.iter_rows():
                    for cell in row:
                      if isinstance(cell.value, str) and cell.value.endswith('.pdf'):
                        cell.value = cell.value[:-4]            

        wb.save(output_path)


@app.route('/')
def index():
    return render_template('index 3.html')

@app.route('/upload', methods=['POST'])
def upload_files():
    registration_numbers = request.form.get('regNumbers')
    uploaded_files = request.files.getlist('pdfFiles')

    if not registration_numbers:
        return jsonify({'success': False, 'message': 'No registration numbers provided.'})

    if not uploaded_files:
        return jsonify({'success': False, 'message': 'No files uploaded.'})

    response = {
        'success': True,
        'message': 'Files uploaded successfully.'
    }

    return jsonify(response)

@app.route('/download/<filename>')
def download_file(filename):
    results_dir = './results'
    file_path = os.path.join(results_dir, filename)
    if os.path.exists(file_path):
        return send_file(file_path, as_attachment=True)
    else:
        return "File not found", 404

@app.route('/generate_cam', methods=['POST'])
def generate_cam():
    data = request.get_json()
    file_names = data.get('fileNames', [])

    cam_table = []
    for pdf_name in file_names:
        row = [pdf_name[:-4]] + ['' for _ in range(7)]  # Remove .pdf extension
        cam_table.append(row)

    return jsonify({'cam_table': cam_table}) 

@app.route('/calculateAttainment', methods=['POST'])
def calculate_attainment_route():
    try:
        reg_numbers = request.form.get('regNumbers')
        components_array = request.form.get('ComponentsArray')
        file_order = request.form.get('fileOrder')
        if not reg_numbers or not components_array:
            return jsonify({'error': 'Missing regNumbers or ComponentsArray'})

        reg_numbers = json.loads(reg_numbers)
        components_array = json.loads(components_array)
        file_order = json.loads(file_order)
        pdf_files = request.files.getlist('pdfFiles')
        uploaded_files = pdf_files

        upload_dir = './uploads'
        results_dir = './results'
        if not os.path.exists(upload_dir):
            os.makedirs(upload_dir)
        if not os.path.exists(results_dir):
            os.makedirs(results_dir)

        file_info = []
        for file in uploaded_files:
            file_path = os.path.join(upload_dir, file.filename)
            try:
                file.save(file_path)
                saved_file_size = os.path.getsize(file_path)
                file_info.append({'name': file.filename[:-4], 'size': saved_file_size})  # Remove .pdf extension
            except Exception as e:
                return jsonify({'success': False, 'message': f'Error saving file {file.filename}'})

        results, info = process_pdfs_in_folder(file_order, upload_dir, reg_numbers)
        df = pd.DataFrame(results).fillna('Not Found')
        df_transposed = df.transpose()
        components = df_transposed.columns.tolist()
        split_up_df = input_split_up(components, components_array)
        total_attainment = calculate_total_possible_attainment(split_up_df)

        split_up_df = calculate_row_wise_sum(split_up_df)
        split_up_df_without_sum = split_up_df.drop(columns=['Row-wise Sum'])

        attainment_table = calculate_attainment(split_up_df, df_transposed, reg_numbers)

        df_transposed = df_transposed.round(2)
        split_up_df_without_sum = split_up_df_without_sum.round(2)
        attainment_table = attainment_table.round(2)

        output_filename = 'results.xlsx'
        output_path = os.path.join(results_dir, output_filename)

        write_to_excel(df_transposed, split_up_df_without_sum, attainment_table, total_attainment, [0.75 * total_attainment[co] for co in ['CO1', 'CO2', 'CO3', 'CO4', 'CO5', 'CO6']], output_path, info, file_order)

        response = {
            'success': True,
            'message': 'Attainments are calculated successfully.',
            'files': file_info,
            'download_url': f'/download/{output_filename}'
        }

        for file in uploaded_files:
            os.remove(os.path.join(upload_dir, file.filename))

        return jsonify(response)
    except Exception as e:
        return jsonify({'error': str(e)}), 500

if __name__ == '__main__':
    app.run(debug=True)
